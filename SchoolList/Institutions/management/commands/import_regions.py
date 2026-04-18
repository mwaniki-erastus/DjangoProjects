from pathlib import Path
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from Institutions.models import Region

class Command(BaseCommand):
    def handle(self, *args, **kwargs):
        BASE_DIR = Path(__file__).resolve().parents[3]
        file_path = BASE_DIR / "data" / "regions.xlsx"

        wb = load_workbook(filename=file_path)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2,values_only=True):
            region_name, region_code

            Region.objects.create(
                region_name = region_name,
                region_code = region_code,
            )

        self.stdout.write(self.style.SUCCESS("Regions imported successfully from XLSX"))

